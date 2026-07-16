"""
911文档订阅添加插件
每天21:01自动扫描金山文档，将文档中「本月更新【国产剧】」「本月更新【国外剧】」「本月更新【综艺】」
三个区段内的在播新剧自动添加到MP订阅（跳过已完结的旧季大包）
"""
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Any, List, Dict, Tuple, Optional

import requests
import openpyxl
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from app.log import logger
from app.plugins import _PluginBase
from app.chain.subscribe import SubscribeChain
from app.schemas.types import MediaType, NotificationType
from app.db.subscribe_oper import Subscribe, SubscribeOper
from app.db import SessionFactory


class Doc911Subscribe(_PluginBase):
    """911文档订阅添加插件"""

    plugin_name = "911文档订阅添加"
    plugin_desc = (
        "每天21:01扫描金山文档，自动添加「本月更新【国产剧】」"
        "「本月更新【国外剧】」「本月更新【综艺】」在播新剧到MP订阅。"
    )
    plugin_icon = "https://raw.githubusercontent.com/jxxghp/MoviePilot-Plugins/main/icons/chain.png"
    plugin_version = "1.0.01"
    plugin_author = "jinyuhao-886"
    plugin_priority = 10

    # 中文数字→整数映射（覆盖1~30，够用了）
    _CN_NUMS = {
        "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
        "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
        "十一": 11, "十二": 12, "十三": 13, "十四": 14, "十五": 15,
        "十六": 16, "十七": 17, "十八": 18, "十九": 19, "二十": 20,
        "二十一": 21, "二十二": 22, "二十三": 23, "二十四": 24, "二十五": 25,
        "二十六": 26, "二十七": 27, "二十八": 28, "二十九": 29, "三十": 30,
    }

    _enabled = False
    _cookies = ""
    _file_id = "206871482430"
    _only_once = False
    _name_aliases = {}  # 原名 → {name, season}
    _scheduler = None

    # 感兴趣的区段
    _SECTIONS = {
        "本月更新【国产剧】": "domestic",
        "本月更新【国外剧】": "foreign",
        "本月更新【综艺】": "variety",
    }
    # 终止区段（遇到这些就停止解析）
    _STOP_SECTIONS = [
        "近期完结剧集",
        "本月更新电影",
        "本月更新【动漫】",
        "本月更新【国漫】",
        "本月更新【日漫】",
        "本月更新【动画】",
        "本月更新【纪录片】",
        "高清经典回看",
        "经典回看",
        "蓝光原盘",
    ]

    def init_plugin(self, config: dict = None):
        """初始化插件"""
        if config:
            self._enabled = config.get("enabled", False)
            self._cookies = config.get("cookies", "")
            self._doc_url = config.get("doc_url", "https://www.kdocs.cn/l/cgqWkD4v1nHK")
            self._file_id = config.get("file_id", "206871482430")
            self._only_once = config.get("only_once", False)
            # 解析别名映射表
            self._name_aliases = self._parse_aliases(
                config.get("name_aliases", "")
            )

        self.stop_service()

        if self._enabled and self._cookies:
            logger.info("911文档订阅添加插件已启用，设置每日21:01定时任务")
            self._scheduler = BackgroundScheduler(timezone="Asia/Shanghai")
            self._scheduler.add_job(
                self.sync_from_doc,
                trigger=CronTrigger(hour=21, minute=1, timezone="Asia/Shanghai"),
                name="911文档订阅添加",
            )
            self._scheduler.start()

            if self._only_once:
                logger.info("911文档订阅添加：仅执行一次模式")
                self.sync_from_doc()
        else:
            logger.info(
                "911文档订阅添加插件未启用或缺少Cookie"
                if not self._enabled
                else "911文档订阅添加插件已禁用"
            )

    def sync_from_doc(self):
        """同步文档中的在播剧集到MP订阅"""
        if not self._cookies:
            logger.error("911文档订阅添加：缺少金山文档Cookie，跳过执行")
            return

        logger.info("911文档订阅添加：开始同步文档数据...")

        try:
            # Step 1: 获取下载链接
            download_url = self._get_download_url()
            if not download_url:
                return

            # Step 2: 下载xlsx
            xlsx_data = self._download_xlsx(download_url)
            if not xlsx_data:
                return

            # Step 3: 保存到插件目录（每次覆盖，只保留最新）
            plugin_dir = self.get_data_path()
            xlsx_path = plugin_dir / "doc.xlsx"
            xlsx_path.write_bytes(xlsx_data)
            logger.info(f"911文档订阅添加：文档已保存到 {xlsx_path}")

            # Step 4: 解析在播剧集（过滤掉已完结的旧季）
            shows = self._parse_shows(xlsx_path)
            if not shows:
                logger.info("911文档订阅添加：未找到新的在播剧集")
                return

            logger.info(
                f"911文档订阅添加：找到 {len(shows)} 个在播剧集，开始添加订阅..."
            )

            # Step 4: 逐个添加订阅
            added = 0
            skipped = 0
            for show in shows:
                try:
                    result = self._add_subscription(show)
                    if result:
                        added += 1
                    else:
                        skipped += 1
                except Exception as e:
                    logger.error(
                        f"911文档订阅添加：添加订阅失败 [{show['name']}]: {str(e)}"
                    )
                    skipped += 1

            logger.info(
                f"911文档订阅添加：执行完成，新增 {added} 个订阅，跳过 {skipped} 个"
            )

        except Exception as e:
            logger.error(f"911文档订阅添加：同步失败: {str(e)}")

    def _get_download_url(self) -> Optional[str]:
        """获取金山文档下载链接"""
        url = (
            f"https://www.kdocs.cn/api/v3/office/file/{self._file_id}/download"
        )
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/130.0.0.0 Safari/537.36"
            ),
            "Referer": self._doc_url,
            "Cookie": self._cookies,
        }
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code != 200:
                logger.error(
                    f"911文档订阅添加：获取下载链接失败 HTTP {resp.status_code}"
                )
                return None
            data = resp.json()
            download_url = data.get("download_url") or data.get("url")
            if not download_url:
                logger.error(
                    f"911文档订阅添加：下载链接为空: {resp.text[:200]}"
                )
                return None
            return download_url
        except Exception as e:
            logger.error(f"911文档订阅添加：获取下载链接异常: {str(e)}")
            return None

    def _download_xlsx(self, download_url: str) -> Optional[bytes]:
        """下载xlsx文件"""
        try:
            resp = requests.get(download_url, timeout=30)
            if resp.status_code != 200:
                logger.error(
                    f"911文档订阅添加：下载xlsx失败 HTTP {resp.status_code}"
                )
                return None
            return resp.content
        except Exception as e:
            logger.error(f"911文档订阅添加：下载xlsx异常: {str(e)}")
            return None

    def _is_ongoing(self, cell_text: str) -> bool:
        """
        判断是否为在播剧集（而非已完结的整季包）
        
        在播特征：单元格包含 "更新XX" 字样
        已完结特征：只有 "XX集全" 或 "XX期全" 但无 "更新"
        """
        return "更新" in cell_text

    def _parse_shows(self, xlsx_path: Path) -> List[Dict]:
        """
        解析xlsx，提取三个区段中的在播剧集
        
        区段:
          - 本月更新【国产剧】→ domestic
          - 本月更新【国外剧】→ foreign  
          - 本月更新【综艺】→ variety
        
        只采集含有「更新」字样的在播条目，跳过已完结整季包
        
        Returns:
            [{"name": str, "year": str, "section": str}, ...]
        """
        shows = []
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            if not wb.sheetnames:
                logger.warning("911文档订阅添加：xlsx中没有sheet")
                return shows

            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            logger.info(f"911文档订阅添加：解析sheet [{sheet_name}]")

            # 状态机：0=找区段, non-zero=在某个区段内
            current_section = None

            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row or 500, values_only=True
            ):
                col_b = row[1] if len(row) > 1 else None
                if not col_b or not isinstance(col_b, str):
                    continue

                cell_text = col_b.strip()
                if not cell_text:
                    continue

                # === 检测目标区段 header ===
                found_section = None
                for keyword, sect_name in self._SECTIONS.items():
                    if keyword in cell_text:
                        found_section = sect_name
                        break

                if found_section:
                    current_section = found_section
                    logger.debug(
                        f"911文档订阅添加：进入区段 [{found_section}]"
                    )
                    continue

                # === 检测终止区段（遇到就停） ===
                if current_section and any(
                    kw in cell_text for kw in self._STOP_SECTIONS
                ):
                    logger.debug(
                        f"911文档订阅添加：遇到终止区段 [{cell_text[:30]}]，停止"
                    )
                    current_section = None
                    continue

                # 不在目标区段，跳过
                if not current_section:
                    continue

                # 数据行必须以 【 开头
                if not cell_text.startswith("【"):
                    continue

                # === 过滤：只在播的（有「更新」字样），跳过已完结整季包 ===
                if not self._is_ongoing(cell_text):
                    logger.debug(
                        f"911文档订阅添加：跳过已完结 [{cell_text[:50]}]"
                    )
                    continue

                # 提取剧名和年份
                show_info = self._extract_show_info(cell_text)
                if show_info:
                    show_info["section"] = current_section
                    shows.append(show_info)
                    logger.debug(
                        f"911文档订阅添加：发现在播剧集 "
                        f"[{show_info['name']} ({show_info['year']})] "
                        f"[{current_section}]"
                    )

            wb.close()
            logger.info(
                f"911文档订阅添加：解析完成，共找到 {len(shows)} 个在播剧集"
            )
            return shows

        except Exception as e:
            logger.error(f"911文档订阅添加：解析xlsx异常: {str(e)}")
            return shows

    def _extract_show_info(self, cell_text: str) -> Optional[Dict]:
        """
        从单元格文本中提取剧名、年份和季数

        格式示例:
            【4K国产剧】百花杀 (2026)[爱情 古装][孟子义].2160P.共36集全更新04
            【HD韩剧】金特务：本色回归 김부장 (2026)[动作][苏志燮 崔大勋].1080P.
            【4K美剧】龙之家族 第三季 (2026)[动作 奇幻 冒险].2160P.HDR.DV.共8集.更新03
            【4K综艺】地球超新鲜 第二季 (2026)[真人秀].1080P+4K.更新7-5.第二期下
            【4K综艺】中餐厅 第十季 (2026)[真人秀].2160P.共12期.第3期.更新7-4

        Returns:
            {"name": str, "year": str, "season": Optional[int]}
        """
        # 匹配 】和 (年份) 之间的剧名
        match = re.search(r"】(.+?) \((\d{4})\)", cell_text)
        if match:
            name = match.group(1).strip()
            year = match.group(2)

            # 清理名字：去掉末尾可能的特殊字符
            name = re.sub(r"[\[\]\.\,\s]+$", "", name)

            if name and year and len(year) == 4:
                result = {"name": name, "year": year, "season": None}

                # 提取季数（按优先级依次尝试，匹配后去掉季数描述再返回）
                season = None
                season_end = None

                # 1. 第N季（阿拉伯数字）：如 "第1季" "第2季"
                m = re.search(r"第(\d+)季$", name)
                if m:
                    season = int(m.group(1))
                    season_end = m.start()

                # 2. 第N季（中文数字）：如 "第一季" "第三季" "第十季"
                if season is None:
                    m = re.search(r"第([一二三四五六七八九十百]+)季$", name)
                    if m:
                        _cn = m.group(1)
                        season = self._CN_NUMS.get(_cn)
                        season_end = m.start()

                # 3. S01 / S1 格式：如 "S01" "S2" "S01E01"
                if season is None:
                    m = re.search(r"[Ss](\d{1,2})(?:[Ee]\d+)?$", name)
                    if m:
                        season = int(m.group(1))
                        season_end = m.start()

                # 4. Season X 格式：如 "Season 1" "Season 01"
                if season is None:
                    m = re.search(r"[Ss]eason\s*(\d{1,2})$", name)
                    if m:
                        season = int(m.group(1))
                        season_end = m.start()

                if season is not None and season_end is not None:
                    result["season"] = season
                    # 从名字中去掉季数描述（如 " 第三季" " S01"）
                    result["name"] = name[:season_end].strip()

                return result

        return None

    def _add_subscription(self, show: Dict) -> bool:
        """
        添加单个剧集订阅
        
        流程：
        1. 通过 SubscribeChain.add() 直接订阅（内部已集成 CustomIdentifiers + TMDB 搜索）
        2. 识别失败则跳过，宁可不订也不乱订
        
        Returns True=已订阅, False=跳过(已存在/识别失败)
        """
        # 检查是否已订阅（按名称匹配）
        db = SessionFactory()
        base_name = show["name"].split(" 第")[0] if " 第" in show["name"] else show["name"]
        existing = Subscribe.get_by_title(db, base_name)
        if existing:
            logger.info(
                f"911文档订阅添加：[{show['name']} ({show['year']})] 已订阅，跳过"
            )
            db.close()
            return False

        try:
            # 通过 SubscribeChain.add() 识别并订阅
            # 内部流程：CustomIdentifiers(自定义识别词) → TMDB 搜索 → 添加订阅
            chain = SubscribeChain()
            add_kwargs = dict(
                title=show["name"],
                season=show.get("season"),
                mtype=MediaType.TV,
                source="911文档订阅添加",
                message=True,
            )
            # 有季数时（如"第三季"）：不传 year，避免 TMDB 按首播年份过滤导致搜不到
            # 无季数时（新剧）：传 year 辅助 TMDB 更精确匹配
            if not show.get("season") and show.get("year"):
                add_kwargs["year"] = show["year"]
            subscribe_id, msg = chain.add(**add_kwargs)

            if subscribe_id:
                logger.info(
                    f"911文档订阅添加：已订阅 "
                    f"[{show['name']} ({show['year']})] "
                    f"subscribe_id={subscribe_id}"
                )
                db.close()
                return True
            
            # 识别失败，发送通知提醒
            logger.warning(
                f"911文档订阅添加：识别失败，跳过 [{show['name']} ({show['year']})]: {msg}"
            )
            try:
                self.post_message(
                    mtype=NotificationType.Subscribe,
                    title=f"新增订阅未能识别 ⚠️",
                    text=f"剧名：{show['name']} ({show['year']})\n"
                         f"类型：{show.get('section', '剧集')}\n"
                         f"原因：{msg}\n"
                         f"建议：可手动搜索TMDB或添加自定义识别词",
                )
            except Exception as e:
                logger.error(f"911文档订阅添加：发送通知失败: {e}")
            db.close()
            return False
            
        except Exception as e:
            logger.error(
                f"911文档订阅添加：添加订阅异常 [{show['name']}]: {str(e)}"
            )
            db.close()
            return False

    def _call_agent_for_show(self, show: Dict):
        """
        调用 MP 智能助手（AI Agent）辅助搜索和订阅
        
        利用 AI Agent 的 SearchMedia + AddSubscribe 工具，
        让 AI 自己识别正确的 TMDB 条目后订阅
        """
        section_names = {
            "domestic": "国产剧",
            "foreign": "国外剧",
            "variety": "综艺",
        }
        section_label = section_names.get(
            show.get("section", ""), "剧集"
        )
        
        prompt = (
            f"请帮我添加一个订阅：\n"
            f"剧名：{show['name']}\n"
            f"年份：{show['year']}\n"
            f"类型：{section_label}（电视剧）\n"
            f"\n"
            f"注意：文档中的剧名可能不是TMDB标准名称，"
            f"请先用 search_media 工具搜索正确的TMDB条目，"
            f"确认后再用 add_subscribe 工具添加订阅。"
            f"如果是综艺节目，请搜索正确的综艺节目名称。"
        )
        
        logger.info(
            f"911文档订阅添加：调用智能助手处理 "
            f"[{show['name']} ({show['year']})]"
        )
        
        # 在调度器线程中运行 async 调用
        import asyncio
        from app.agent import AgentManager
        
        try:
            asyncio.run(
                AgentManager.run_background_prompt(
                    message=prompt,
                    session_prefix="__doc911_agent",
                )
            )
            logger.info(
                f"911文档订阅添加：智能助手已接收任务 "
                f"[{show['name']}]"
            )
        except Exception as e:
            logger.error(
                f"911文档订阅添加：调用智能助手失败 "
                f"[{show['name']}]: {str(e)}"
            )

    def _parse_aliases(self, raw: str) -> Dict[str, dict]:
        """
        解析别名映射配置
        
        格式：每行一个映射
          原名=TMDB标准名
          原名=TMDB标准名|季数
        
        示例：
          歌手=我是歌手|11
          乘风=乘风破浪的姐姐
        """
        aliases = {}
        if not raw or not raw.strip():
            return aliases
        
        for line in raw.strip().split("\n"):
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("//"):
                continue
            if "=" not in line:
                continue
            
            parts = line.split("=", 1)
            src_name = parts[0].strip()
            target = parts[1].strip()
            
            if not src_name or not target:
                continue
            
            entry = {"name": target}
            if "|" in target:
                target_parts = target.split("|", 1)
                entry["name"] = target_parts[0].strip()
                season_str = target_parts[1].strip()
                try:
                    entry["season"] = int(season_str)
                except ValueError:
                    pass
            
            aliases[src_name] = entry
            logger.info(
                f"911文档订阅添加：别名映射 [{src_name}] → "
                f"{entry['name']}" + 
                (f" S{entry['season']}" if "season" in entry else "")
            )
        
        return aliases

    def _apply_alias(self, show: Dict) -> Dict:
        """
        对剧集信息应用别名映射
        
        如果原名在映射表中，替换为TMDB标准名
        """
        name = show["name"]
        for src_name, target in self._name_aliases.items():
            if name == src_name or name.startswith(src_name + " "):
                show["name"] = target["name"]
                if "season" in target:
                    show["season"] = target["season"]
                logger.info(
                    f"911文档订阅添加：别名映射生效 "
                    f"[{name}] → [{target['name']}]"
                    + (f" S{target['season']}" if "season" in target else "")
                )
                break
        return show

    def get_state(self) -> bool:
        """获取插件状态"""
        return self._enabled and bool(self._cookies)

    def get_form(self) -> Tuple[List[dict], Dict[str, Any]]:
        """获取插件配置表单"""
        return [
            {
                "component": "VForm",
                "content": [
                    {
                        "component": "VRow",
                        "content": [
                            {
                                "component": "VCol",
                                "props": {"cols": 12, "md": 6},
                                "content": [
                                    {
                                        "component": "VSwitch",
                                        "props": {
                                            "model": "enabled",
                                            "label": "启用插件",
                                        },
                                    }
                                ],
                            },
                            {
                                "component": "VCol",
                                "props": {"cols": 12, "md": 6},
                                "content": [
                                    {
                                        "component": "VSwitch",
                                        "props": {
                                            "model": "only_once",
                                            "label": "保存后立即执行一次",
                                            "hint": "勾选后保存配置时立即扫描一次",
                                        },
                                    }
                                ],
                            },
                        ],
                    },
                    {
                        "component": "VRow",
                        "content": [
                            {
                                "component": "VCol",
                                "props": {"cols": 12},
                                "content": [
                                    {
                                        "component": "VTextarea",
                                        "props": {
                                            "model": "cookies",
                                            "label": "金山文档Cookie",
                                            "placeholder": "完整的Cookie字符串，含 wps_sid=...; csrf=...; uid=...",
                                            "rows": 4,
                                            "hint": "登录 https://www.kdocs.cn 后F12获取。Cookie过期后需更新。",
                                        },
                                    }
                                ],
                            }
                        ],
                    },
                    {
                        "component": "VRow",
                        "content": [
                            {
                                "component": "VCol",
                                "props": {"cols": 12, "md": 12},
                                "content": [
                                    {
                                        "component": "VTextField",
                                        "props": {
                                            "model": "doc_url",
                                            "label": "911文档链接",
                                            "placeholder": "https://www.kdocs.cn/l/cgqWkD4v1nHK",
                                            "hint": "金山文档的分享链接。更换文档时填入新链接即可，无需手动修改File ID。",
                                            "clearable": True,
                                        },
                                    }
                                ],
                            }
                        ],
                    },
                    {
                        "component": "VRow",
                        "content": [
                            {
                                "component": "VCol",
                                "props": {"cols": 12},
                                "content": [
                                    {
                                        "component": "VAlert",
                                        "props": {
                                            "type": "info",
                                            "text": (
                                                "📋 每天21:01定时扫描文档中三个区段：\n"
                                                "• 本月更新【国产剧】\n"
                                                "• 本月更新【国外剧】\n"
                                                "• 本月更新【综艺】\n"
                                                "自动过滤已完结整季包，只订阅在播新剧。\n"
                                                "TMDB识别失败的条目会发送通知提醒，不会自动补填。"
                                            ),
                                        },
                                    }
                                ],
                            }
                        ],
                    },
                ],
            }
        ], {
            "enabled": False,
            "only_once": False,
            "cookies": "",
            "doc_url": "https://www.kdocs.cn/l/cgqWkD4v1nHK",
        }
    def get_page(self) -> List[dict]:
        """获取插件页面"""
        pass

    def get_api(self) -> List[Dict[str, Any]]:
        """注册插件API"""
        return []

    def stop_service(self):
        """停止插件服务"""
        if self._scheduler:
            try:
                self._scheduler.shutdown(wait=False)
            except Exception as e:
                logger.error(
                    f"911文档订阅添加：停止调度器失败: {str(e)}"
                )
            self._scheduler = None
